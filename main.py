import os
import locale
import PyPDF2
import shutil
from unidecode import unidecode
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def sort_dir_ascending(dir_path):
    files = os.listdir(dir_path)
    sorted_files = sorted(files)
    paths_sorted_files = [os.path.join(dir_path, file) for file in sorted_files]
    return paths_sorted_files

# Extract the study and patient name from PDF. 
def get_result_info(file_path):
    with open(file_path, 'rb') as file:
        reader = PyPDF2.PdfReader(file)
        content = reader.pages[0]
        text = content.extract_text()
        words = text.split()
        patients_name = ''
        study_name = ''

        try:
            # Extract the study name between the words 'ESTUDIO' y 'HALLAZGOS:' 
            start = words.index('ESTUDIO') + 2
            end = words.index('HALLAZGOS:')
            
            for study_words in range(start, end):
                study_name += f'{words[study_words]} '

            # Extract the patients name between the words 'NOMBRE' y 'DOCUMENTO' 
            start = words.index('NOMBRE') + 2
            end = words.index('DOCUMENTO')
            
            for study_words in range(start, end):
                patients_name += f'{words[study_words]} '

            return (study_name.strip(), patients_name.strip()) 
        except ValueError:
            return (ValueError, ValueError)

def change_background_color(cell, color):
    color = PatternFill(patternType='solid', fgColor=color)
    cell.fill = color

def remove_accents(text): 
    word_without_accents = ''
    
    for letter in text:
        if letter != 'ñ' and letter != 'Ñ':
            word_without_accents += unidecode(letter)
        else:
            word_without_accents += letter      
    return word_without_accents

def check_duplicate_items(list):
    seen = set()
    duplicates = []

    for item in list:
        if item in seen:
            duplicates.append(item)

        seen.add(item)
    return duplicates

def check_files_in_report(spreadsheet_path, sheet, 
                          patients_column, study_column, 
                          files_path, invalid_files_path): 
    study_patients_folder = []
    study_patients_sheet = []

    # Extract data from folder.
    for file in os.listdir(files_path):
        if file != 'desktop.ini':
            file_path = f'{files_path}\\{file}'
            file_study, file_patient = get_result_info(file_path)

            # Move files with invalid format.
            if file_study == ValueError:
                os.makedirs(invalid_files_path, exist_ok=True)
                shutil.move(file_path, invalid_files_path)
            else: 
                file_patient = remove_accents(file_patient)
                file_study = remove_accents(file_study)
                patient_info = f'{file_patient.strip()} _ {file_study.strip()}'.lower()

                study_patients_folder.append(patient_info)


    # Extract data from spreadsheet.
    spreadsheet = load_workbook(spreadsheet_path)
    sheet = spreadsheet[sheet]
    green = '92d050'

    for i in range(2, (len(sheet[patients_column]) + 1)):
        sheet_study = remove_accents(sheet[f'{study_column}{i}'].value)
        sheet_patient = remove_accents(sheet[f'{patients_column}{i}'].value)
        patient_info = f"{sheet_patient.strip()} _ {sheet_study.strip()}".lower()
        study_patients_sheet.append(patient_info)

    files_in_report = set(study_patients_folder).intersection(set(study_patients_sheet))
    locale.setlocale(locale.LC_ALL, 'Spanish_Spain.1252')
    files_in_report = sorted(list(files_in_report), key=locale.strxfrm)
    
    
    # Highlight patients cell who have PDF in folder. 
    i, j = 0, 2
    while (i < len(files_in_report)) and (j < (len(sheet[patients_column]) + 1)):
        cell_study = remove_accents(sheet[f'{study_column}{j}'].value)
        cell_patient = remove_accents(sheet[f'{patients_column}{j}'].value)
        patient_info = f'{cell_patient.strip()} _ {cell_study.strip()}'.lower()
    
        if files_in_report[i] == patient_info:
            change_background_color(sheet[f'{patients_column}{j}'], green)
            i += 1
            j += 1
        else:
            j += 1
    
    try:
        spreadsheet.save(spreadsheet_path)
    except PermissionError:
        print('The changes to the spreadsheet could not be saved because it is open \
              by another program!!!.\n')

    study_patients_dups_sheet = check_duplicate_items(study_patients_sheet)
    study_patients_dups_folder = check_duplicate_items(study_patients_folder)
    files_out_report = set(study_patients_folder).difference(set(study_patients_sheet))

    print(f'FILES OUT OF REPORT:')
    for index, value in enumerate(files_out_report): 
        print(f'{index + 1}) {value}')

    print(F'\nDUPLICATE CELLS:')
    for index, value in enumerate(study_patients_dups_sheet): 
        print(f'{index + 1}) {value}')

    print(F'\nDUPLICATE FILES:')
    for index, value in enumerate(study_patients_dups_folder): 
        print(f'{index + 1}) {value}')
    

# File directory.
files_path = r'D:\Users\User\Documents\PDF'
spreadsheet_path = r'D:\Users\User\Documents\TRAZABILIDAD.xlsx'
invalid_files_path = r'D:\Users\User\Documents\Archivos con formato inválido'
patients_column = "D"
study_column = "L"
sheet = "Hoja1"

check_files_in_report(spreadsheet_path, sheet, 
                      patients_column, study_column, 
                      files_path, invalid_files_path)